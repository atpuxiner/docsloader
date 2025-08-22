import binascii
import logging
from typing import AsyncGenerator

import xlrd
from openpyxl.reader.excel import load_workbook

from docsloader.base import BaseLoader, DocsData

logger = logging.getLogger(__name__)


class XlsxLoader(BaseLoader):

    async def load_by_basic(self) -> AsyncGenerator[DocsData, None]:
        with open(self.tmpfile, "rb") as f:
            header_flag = binascii.hexlify(f.read(8)).decode().upper()
        if header_flag.startswith("504B0304"):  # .xlsx
            wb = load_workbook(filename=self.tmpfile, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.info(f"Processing sheet: {sheet_name}")
                rows = ws.iter_rows(values_only=True)
                try:
                    header = next(rows)
                except StopIteration:
                    header = []
                    rows = []
                idx = 0
                self.metadata.update(
                    header=header,
                    sheet_name=sheet_name,
                )
                for idx, row in enumerate(rows):
                    yield DocsData(
                        idx=idx,
                        type="text",
                        data=row,
                        metadata=self.metadata,
                    )
                if not idx:
                    yield DocsData(
                        idx=0,
                        type="text",
                        data=[],
                        metadata=self.metadata,
                    )
            wb.close()
        elif header_flag.startswith("D0CF11E0A1B11AE1"):  # .xls
            book = xlrd.open_workbook(self.tmpfile, formatting_info=False)
            for sheet_name in book.sheet_names():
                sheet = book.sheet_by_name(sheet_name)
                logger.info(f"Processing sheet: {sheet_name}")
                header = sheet.row_values(0) if sheet.nrows > 0 else []
                self.metadata.update(
                    header=header,
                    sheet_name=sheet_name,
                )
                if sheet.nrows > 1:
                    for idx in range(1, sheet.nrows):
                        row = sheet.row_values(idx)
                        yield DocsData(
                            idx=idx - 1,
                            type="text",
                            data=row,
                            metadata=self.metadata,
                        )
                else:
                    yield DocsData(
                        idx=0,
                        type="text",
                        data=[],
                        metadata=self.metadata,
                    )
        else:
            raise ValueError(f"Unsupported file format: {header_flag}")
